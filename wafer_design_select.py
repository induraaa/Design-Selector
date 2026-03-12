"""
Wafer Design Picker Tool  v1
Universal format detection — supports all known SCR file variants.

Detected formats
────────────────
  A  null='.'  edge='X'  designs=1,'1b','1c'…   (SCR05, SCR07, Design_1)
  B  null='-'  edge='X'  designs=1,2,3…          (SCR03, SCR04, ARREST, SCR_Multimap)
  C  null='___' edge=N/A designs='001','002'…     (SCR06 — RowData: prefix in col A)

Install:  pip install openpyxl pillow numpy
Run:      python wafer_map_tool.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading, os, re, time
from collections import Counter

def _ensure(pkg, imp=None):
    import importlib, subprocess, sys
    try: importlib.import_module(imp or pkg)
    except ImportError:
        subprocess.check_call([sys.executable,"-m","pip","install",pkg,"-q"])

_ensure("openpyxl"); _ensure("pillow","PIL"); _ensure("numpy")

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.cell import WriteOnlyCell
import numpy as np
from PIL import Image, ImageTk

import math, tempfile, base64


# ── palette ───────────────────────────────────────────────────────────────────
PAL_RGB = [
    (  0,103,192),(234, 67, 53),( 52,168, 83),(251,188,  4),(103, 58,183),
    (  0,150,136),(255,111, 26),( 66,133,244),(139,195, 74),(233, 30, 99),
    (  0,188,212),( 96,125,139),(255,152,  0),( 76,175, 80),(156, 39,176),
    (244, 81, 30),(  0,137,123),(121, 85, 72),(  3,169,244),(175,180, 43),
]
PAL_HEX  = ["#%02x%02x%02x"%c for c in PAL_RGB]
PAL_XLSX = ["%02X%02X%02X"%c for c in PAL_RGB]

SEL_RGB=(0,170,0); SEL_XLSX="00AA00"
X_RGB  =(110,110,110); X_XLSX="6E6E6E"
x_RGB  =(210,210,210)

WIN_BG   = "#f0f0f0"
PANEL_BG = "#f5f5f5"
SEP_COL  = "#d0d0d0"
HDR_BG   = "#e1e1e1"
BLUE     = "#0067c0"
BLUE_HOV = "#005a9e"


# ═══════════════════════════════════════════════════════════════════════
#  FAST XLSX READER  (direct zipfile + XML — no openpyxl for loading)
# ═══════════════════════════════════════════════════════════════════════
import zipfile as _zipfile

_COL_CACHE = {}
def _col_idx(col_str):
    """Convert Excel column letters to 0-based index, cached."""
    if col_str in _COL_CACHE:
        return _COL_CACHE[col_str]
    c = 0
    for ch in col_str:
        c = c * 26 + (ord(ch) - 64)
    v = c - 1
    _COL_CACHE[col_str] = v
    return v

_ROW_PAT  = re.compile(rb'<row\s[^>]*r="(\d+)"[^>]*>(.*?)</row>', re.DOTALL)
_CELL_PAT = re.compile(rb'<c\s+r="([A-Z]+)\d+"(?:[^>]*\bt="([^"]*)")?[^>]*>(?:<v>([^<]*)</v>)?')
_SS_PAT   = re.compile(rb'<t(?:\s[^>]*)?>([^<]*)</t>')
_SH_PAT   = re.compile(rb'<sheet\s+name="([^"]*)"[^/]*/>')

MAP_KW   = ('map', 'design', 'full', 'all')
SKIP_KW  = ('count', 'summary', 'field', 'layout', 'die_count')

def _fast_read_xlsx(filepath):
    """Read the map sheet from an xlsx file directly via zipfile+regex.
    Returns (raw, sheet_name) where raw is list[list[str|None]].
    2-11x faster than openpyxl for large wafer map files."""
    with _zipfile.ZipFile(filepath) as zf:
        names = set(zf.namelist())

        # Shared strings table
        ss = []
        if 'xl/sharedStrings.xml' in names:
            ss_xml = zf.read('xl/sharedStrings.xml')
            ss = [m.group(1).decode('utf-8') for m in _SS_PAT.finditer(ss_xml)]

        # Sheet list from workbook
        wb_xml    = zf.read('xl/workbook.xml')
        sh_names  = [m.group(1).decode('utf-8') for m in _SH_PAT.finditer(wb_xml)]

        # Pick best map sheet by name score
        best_i = max(
            range(len(sh_names)),
            key=lambda i: (
                sum(2 for k in MAP_KW  if k in sh_names[i].lower()) -
                sum(1 for k in SKIP_KW if k in sh_names[i].lower())
            )
        )
        sheet_name = sh_names[best_i]

        sh_xml = zf.read(f'xl/worksheets/sheet{best_i + 1}.xml')

    # Parse cells
    rows_dict = {}
    max_col   = 0

    for rm in _ROW_PAT.finditer(sh_xml):
        rn   = int(rm.group(1))
        rxml = rm.group(2)
        cells = {}
        for cm in _CELL_PAT.finditer(rxml):
            typ   = cm.group(2)
            raw_v = cm.group(3)
            if raw_v is None:
                continue
            raw_v = raw_v.decode('utf-8')
            if typ == b's':
                val = ss[int(raw_v)] if ss else raw_v
            else:
                val = raw_v
            ci = _col_idx(cm.group(1).decode())
            cells[ci] = val
            if ci > max_col:
                max_col = ci
        if cells:
            rows_dict[rn] = cells

    if not rows_dict:
        return [], sheet_name

    max_row = max(rows_dict.keys())
    w       = max_col + 1
    raw = []
    for r in range(1, max_row + 1):
        row = rows_dict.get(r, {})
        raw.append([row.get(c) for c in range(w)])

    return raw, sheet_name


# ═══════════════════════════════════════════════════════════════════════
#  UNIVERSAL FORMAT DETECTOR + PARSER
# ═══════════════════════════════════════════════════════════════════════
def detect_and_parse(filepath):
    """
    Opens an xlsx file, auto-detects its format, and returns:
      grid        : list[list]  — normalised grid (null='.', edge='X', design=str)
      designs     : list[str]   — sorted unique design IDs
      counts      : dict        — {design: die_count}
      sheet_name  : str
      fmt         : 'A' | 'B' | 'C'
      null_char   : str         — original null character found in file
    """
    raw, sheet_name = _fast_read_xlsx(filepath)

    # ── detect format ────────────────────────────────────────────────────
    # Sample up to 5000 non-None values for format detection
    sample = []
    for row in raw:
        for c in row:
            if c is not None:
                sample.append(c)
            if len(sample) >= 5000:
                break
        if len(sample) >= 5000:
            break

    val_counts = Counter(sample)

    has_rowdata   = val_counts.get('RowData:', 0) > 10 or \
                    any(str(v).startswith('RowData:') for v in sample[:50])
    has_dash_null = val_counts.get('-', 0) > 100

    if has_rowdata:
        fmt = 'C'
    elif has_dash_null:
        fmt = 'B'
    else:
        fmt = 'A'

    # ── parse into normalised grid ───────────────────────────────────────
    if fmt == 'C':
        grid, null_char = _parse_format_c(raw)
    elif fmt == 'B':
        grid = _normalise_grid(raw, null_chars={'-'}, edge_chars={'X'})
        null_char = '-'
    else:
        grid = _normalise_grid(raw, null_chars={'.', None}, edge_chars={'X'})
        null_char = '.'

    # ── collect designs ──────────────────────────────────────────────────
    counts = Counter()
    for row in grid:
        for cell in row:
            if cell not in ('.', 'X', 'x', None):
                counts[str(cell)] += 1

    designs = _sort_designs(list(counts.keys()))

    return grid, designs, dict(counts), sheet_name, fmt, null_char


def _normalise_grid(raw, null_chars, edge_chars):
    """Normalise raw cells: None/null→'.', edge→'X', else str."""
    grid = []
    for row in raw:
        out = []
        for cell in row:
            if cell is None or cell in null_chars:
                out.append('.')
            elif cell in edge_chars:
                out.append('X')
            else:
                out.append(cell)
        grid.append(out)
    return grid


def _parse_format_c(raw):
    """Format C: RowData: prefix, '___' nulls, no edge marker."""
    grid = []
    for row in raw:
        if not row or row[0] != 'RowData:':
            continue
        out = []
        for cell in row[1:]:
            if cell is None or cell == '___':
                out.append('.')
            elif cell:
                out.append(cell.strip())
        if out:
            grid.append(out)
    return grid, '___'


def _sort_designs(design_list):
    """Sort designs naturally: numeric first, then alpha suffix."""
    def key(d):
        m = re.match(r'^(\d+)(.*)$', str(d))
        if m:
            return (int(m.group(1)), m.group(2))
        if str(d).isdigit():
            return (int(d), '')
        return (999999, str(d))
    return sorted(design_list, key=key)


# ═══════════════════════════════════════════════════════════════════════
#  MAIN APPLICATION
# ═══════════════════════════════════════════════════════════════════════
class WaferMapTool(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Design Picker")
        self.minsize(1200, 700)
        self.configure(bg=WIN_BG)
        self.state("zoomed")   # maximised on Windows; harmless on others

        # App icon — load from same folder as script
        try:
            _ico_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "paint.ico")
            self.iconbitmap(_ico_path)
        except Exception:
            pass

        style = ttk.Style(self)
        for t in ("winnative","vista","clam","default"):
            try: style.theme_use(t); break
            except: pass

        # state
        self.filepath         = None
        self.source_sheet     = None
        self.file_fmt         = None
        self.null_char        = '.'
        self.grid_data        = []
        self.enc              = None
        self.designs          = []
        self.design_counts    = {}
        self.selected_designs = set()
        self.cm_rgb           = {}
        self.cm_xlsx          = {}
        self.design_vars      = {}
        self._hover_jobs      = {}
        self.zoom             = 1.0
        self._photo           = None
        self._pending         = False
        self.prev_zoom        = 1.0   # zoom for preview canvas
        self._prev_photo      = None  # keep reference
        self._lut             = None  # colour lookup table (map)
        self._prev_lut        = None  # colour lookup table (preview)
        self._lut_dirty       = True  # rebuild LUT on next render

        self._build_menu()
        self._build_toolbar()
        self._build_body()
        self._build_bottom_bar()
        self.bind("<Control-o>", lambda e: self._open_file())
        self.bind("<Control-e>", lambda e: self._export())

    # ── menu ─────────────────────────────────────────────────────────────
    def _build_menu(self):
        mb = tk.Menu(self); self.config(menu=mb)
        fm = tk.Menu(mb, tearoff=0)
        fm.add_command(label="Open File...\tCtrl+O", command=self._open_file)
        fm.add_separator()
        fm.add_command(label="Export...\tCtrl+E",    command=self._export)
        fm.add_separator()
        fm.add_command(label="Exit",                  command=self.destroy)
        mb.add_cascade(label="File", menu=fm)
        hm = tk.Menu(mb, tearoff=0)
        hm.add_command(label="About", command=lambda: messagebox.showinfo(
            "About",
            "Design Picker\n\n"
            "Supports all SCR wafer map formats:\n"
            "  Format A  —  null='.', edge='X'\n"
            "  Format B  —  null='-', edge='X'\n"
            "  Format C  —  RowData: / null='___'\n\n"
            "Python · Tkinter · NumPy · Pillow · openpyxl"))
        mb.add_cascade(label="Help", menu=hm)

    # ── toolbar ───────────────────────────────────────────────────────────
    def _build_toolbar(self):
        tb = tk.Frame(self, bg=WIN_BG, relief="raised", bd=1)
        tb.pack(fill="x", side="top")

        def btn(text, cmd, w=None, sep_after=False, tag=None):
            b = tk.Button(tb, text=text, command=cmd,
                          relief="flat", bd=0, bg=WIN_BG,
                          font=("Segoe UI", 9), padx=10, pady=5,
                          cursor="hand2", fg="#111")
            if w:
                b.config(width=w)
            # Smooth hover: blue tint in, fade out
            def _hov_in(e, _b=b):
                if str(_b.cget("state")) != "disabled":
                    _b.config(bg="#d9e8f5", fg="#0055aa", relief="groove")
            def _hov_out(e, _b=b):
                if str(_b.cget("state")) != "disabled":
                    _b.config(bg=WIN_BG, fg="#111", relief="flat")
            b.bind("<Enter>", _hov_in)
            b.bind("<Leave>", _hov_out)
            b.pack(side="left", padx=1, pady=2)
            if sep_after:
                tk.Frame(tb, bg=SEP_COL, width=1).pack(
                    side="left", fill="y", pady=4, padx=3)
            if tag:
                setattr(self, tag, b)
            return b

        btn("📂  Open",      self._open_file,  sep_after=True)
        btn("⟳  Refresh",   self._do_convert)
        btn("💾  Export",    self._export,      sep_after=True, tag="tb_export")
        self.tb_export.config(state="disabled", fg="#aaaaaa")

        btn("🔍+  Zoom In",  self._zoom_in)
        btn("🔍−  Zoom Out", self._zoom_out)
        btn("⊞  Fit",        self._zoom_fit,   sep_after=True)

        # Select All / None shortcut buttons in toolbar
        btn("☑  All",        self._select_all)
        btn("☐  None",       self._select_none, sep_after=True)

        # Reset button in toolbar
        btn("↺  Reset",      self._reset)

        self.show_grid = tk.BooleanVar(value=False)
        gc = tk.Checkbutton(tb, text="Grid Lines", variable=self.show_grid,
                       command=self._schedule_redraw,
                       bg=WIN_BG, font=("Segoe UI", 9),
                       relief="flat", cursor="hand2",
                       activebackground="#d9e8f5", selectcolor=WIN_BG)
        gc.pack(side="left", padx=6)
        gc.bind("<Enter>", lambda e: gc.config(bg="#d9e8f5"))
        gc.bind("<Leave>", lambda e: gc.config(bg=WIN_BG))

        self.zoom_lbl = tk.Label(tb, text="Zoom: 100%",
                                  bg=WIN_BG, font=("Segoe UI", 9), fg="#555")
        self.zoom_lbl.pack(side="right", padx=12)

        self.breadcrumb = tk.Label(tb, text="No file loaded",
                                    bg=WIN_BG, font=("Segoe UI", 9), fg="#666")
        self.breadcrumb.pack(side="left", padx=12)

    # ── body ──────────────────────────────────────────────────────────────
    def _build_body(self):
        body = tk.Frame(self, bg=WIN_BG)
        body.pack(fill="both", expand=True, padx=6, pady=(4,0))

        # ── outer PanedWindow: [centre area] | [right panel] ─────────────
        outer_pane = ttk.PanedWindow(body, orient="horizontal")
        outer_pane.pack(fill="both", expand=True)

        # ── centre PanedWindow: [map] | [txt preview] ────────────────────
        centre_pane = ttk.PanedWindow(outer_pane, orient="horizontal")
        outer_pane.add(centre_pane, weight=3)

        # Map card
        map_card = ttk.LabelFrame(centre_pane, text=" Wafer Map ",
                                   labelanchor="nw")
        centre_pane.add(map_card, weight=3)

        self.canvas = tk.Canvas(map_card, bg=WIN_BG,
                                 cursor="crosshair", highlightthickness=0)
        hsc = ttk.Scrollbar(map_card, orient="horizontal", command=self.canvas.xview)
        vsc = ttk.Scrollbar(map_card, orient="vertical",   command=self.canvas.yview)
        self.canvas.configure(xscrollcommand=hsc.set, yscrollcommand=vsc.set)
        hsc.pack(side="bottom", fill="x")
        vsc.pack(side="right",  fill="y")
        self.canvas.pack(fill="both", expand=True, padx=2, pady=2)
        self._draw_placeholder()

        # Bind mouse-wheel zoom: Ctrl+scroll anywhere on canvas
        self.canvas.bind("<MouseWheel>",          self._on_canvas_scroll)
        self.canvas.bind("<Button-4>",            self._on_canvas_scroll)  # Linux scroll up
        self.canvas.bind("<Button-5>",            self._on_canvas_scroll)  # Linux scroll down
        self.canvas.bind("<Control-MouseWheel>",  self._on_canvas_zoom)
        self.canvas.bind("<Control-Button-4>",    self._on_canvas_zoom)
        self.canvas.bind("<Control-Button-5>",    self._on_canvas_zoom)

        # Preview canvas — shows selected design only (same style as map)
        prev_card = ttk.LabelFrame(centre_pane,
                                    text=" Selected Design Preview ",
                                    labelanchor="nw")
        centre_pane.add(prev_card, weight=2)

        self.prev_canvas = tk.Canvas(prev_card, bg="#e8e8e8",
                                      cursor="crosshair", highlightthickness=0)
        ph = ttk.Scrollbar(prev_card, orient="horizontal", command=self.prev_canvas.xview)
        pv = ttk.Scrollbar(prev_card, orient="vertical",   command=self.prev_canvas.yview)
        self.prev_canvas.configure(xscrollcommand=ph.set, yscrollcommand=pv.set)
        ph.pack(side="bottom", fill="x")
        pv.pack(side="right",  fill="y")
        self.prev_canvas.pack(fill="both", expand=True, padx=2, pady=2)
        self._draw_prev_placeholder()

        # Same scroll/zoom bindings as the main canvas
        self.prev_canvas.bind("<MouseWheel>",         self._on_prev_scroll)
        self.prev_canvas.bind("<Button-4>",           self._on_prev_scroll)
        self.prev_canvas.bind("<Button-5>",           self._on_prev_scroll)
        self.prev_canvas.bind("<Control-MouseWheel>", self._on_prev_zoom)
        self.prev_canvas.bind("<Control-Button-4>",   self._on_prev_zoom)
        self.prev_canvas.bind("<Control-Button-5>",   self._on_prev_zoom)

        # Right panel (fixed 270px)
        right = tk.Frame(outer_pane, bg=WIN_BG, width=270)
        right.pack_propagate(False)
        outer_pane.add(right, weight=0)
        self._build_right(right)

    # ── canvas scroll / zoom ──────────────────────────────────────────────
    def _on_canvas_scroll(self, event):
        """Plain scroll → pan vertically (Ctrl held → zoom instead)."""
        if event.state & 0x4:          # Ctrl key is down
            self._on_canvas_zoom(event)
            return
        # pan
        if event.num == 4:
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.canvas.yview_scroll(1, "units")
        else:
            self.canvas.yview_scroll(-1*(event.delta//120), "units")

    def _on_canvas_zoom(self, event):
        """Ctrl+scroll → zoom in/out centred on mouse position."""
        if event.num == 4 or (hasattr(event,'delta') and event.delta > 0):
            factor = 1.15
        else:
            factor = 1/1.15

        old_zoom = self.zoom
        self.zoom = max(0.05, min(16.0, self.zoom * factor))
        if abs(self.zoom - old_zoom) < 0.001:
            return

        # keep the point under the cursor fixed after zoom
        cx = self.canvas.canvasx(event.x)
        cy = self.canvas.canvasy(event.y)
        ratio = self.zoom / old_zoom
        self._schedule_redraw()
        def _adjust():
            sr = self.canvas.cget("scrollregion")
            if not sr: return
            _, _, sw, sh = map(float, sr.split())
            new_cx = cx * ratio
            new_cy = cy * ratio
            vw = self.canvas.winfo_width()
            vh = self.canvas.winfo_height()
            fx = max(0.0, min(1.0, (new_cx - event.x) / sw))
            fy = max(0.0, min(1.0, (new_cy - event.y) / sh))
            self.canvas.xview_moveto(fx)
            self.canvas.yview_moveto(fy)
        self.after(60, _adjust)

    def _build_right(self, parent):
        # ── File Info ────────────────────────────────────────────────────
        fi = ttk.LabelFrame(parent, text=" File Info ", labelanchor="nw")
        fi.pack(fill="x", padx=0, pady=(0,6))

        self.stat_vars = {}
        for lbl, key in [("File","File:"), ("Format","Format:"),
                          ("Grid","Dims:"), ("Total Die","Total Die:"),
                          ("Load Time","Load Time:")]:
            row = tk.Frame(fi, bg=WIN_BG); row.pack(fill="x", padx=8, pady=1)
            tk.Label(row, text=lbl+":", font=("Segoe UI",8),
                     bg=WIN_BG, fg="#666", width=10, anchor="w").pack(side="left")
            v = tk.StringVar(value="—"); self.stat_vars[key] = v
            tk.Label(row, textvariable=v, font=("Segoe UI",8,"bold"),
                     bg=WIN_BG, fg="#111").pack(side="left", fill="x", expand=True)
        tk.Frame(fi, bg=WIN_BG, height=4).pack()

        # ── Designs ──────────────────────────────────────────────────────
        dc = ttk.LabelFrame(parent, text=" Designs ", labelanchor="nw")
        dc.pack(fill="both", expand=True, padx=0, pady=(0,6))

        # column headers
        hdr = tk.Frame(dc, bg=HDR_BG); hdr.pack(fill="x")
        tk.Label(hdr, text="  Design", font=("Segoe UI",8,"bold"),
                 bg=HDR_BG, width=12, anchor="w").pack(side="left", padx=(20,0), pady=3)
        tk.Label(hdr, text="Die Count", font=("Segoe UI",8,"bold"),
                 bg=HDR_BG, anchor="e").pack(side="right", padx=8)
        ttk.Separator(dc).pack(fill="x")

        # scrollable list
        lc = tk.Canvas(dc, bg=WIN_BG, highlightthickness=0, bd=0)
        ls = ttk.Scrollbar(dc, orient="vertical", command=lc.yview)
        lc.configure(yscrollcommand=ls.set)
        ls.pack(side="right", fill="y", padx=(0,2), pady=2)
        lc.pack(side="left", fill="both", expand=True)
        lc.bind("<MouseWheel>",
            lambda e: lc.yview_scroll(-1*(e.delta//120),"units"))
        self._design_canvas = lc

        self.design_inner = tk.Frame(lc, bg=WIN_BG)
        lc.create_window((0,0), window=self.design_inner, anchor="nw")
        self.design_inner.bind("<Configure>",
            lambda e: lc.configure(scrollregion=lc.bbox("all")))

        self._empty_lbl = tk.Label(self.design_inner,
                                    text="Open a file to detect designs",
                                    font=("Segoe UI",8,"italic"),
                                    bg=WIN_BG, fg="#999")
        self._empty_lbl.pack(pady=14)

        ttk.Separator(dc).pack(fill="x")

        # All / None + count
        bot = tk.Frame(dc, bg=WIN_BG); bot.pack(fill="x", padx=8, pady=5)
        tk.Button(bot, text="All",  command=self._select_all,
                  font=("Segoe UI",8), relief="raised", bd=2,
                  padx=6, pady=1).pack(side="left", padx=(0,4))
        tk.Button(bot, text="None", command=self._select_none,
                  font=("Segoe UI",8), relief="raised", bd=2,
                  padx=6, pady=1).pack(side="left")
        self.sel_count_lbl = tk.Label(bot, text="0 selected",
                                       font=("Segoe UI",8),
                                       bg=WIN_BG, fg="#555")
        self.sel_count_lbl.pack(side="right")

        # ── Selection Summary ─────────────────────────────────────────────
        sc = ttk.LabelFrame(parent, text=" Selection ", labelanchor="nw")
        sc.pack(fill="x", padx=0, pady=(0,6))

        r1 = tk.Frame(sc, bg=WIN_BG); r1.pack(fill="x", padx=8, pady=(5,1))
        tk.Label(r1, text="Selected die:", font=("Segoe UI",8),
                 bg=WIN_BG, fg="#666", width=13, anchor="w").pack(side="left")
        self.sel_die_var = tk.StringVar(value="—")
        tk.Label(r1, textvariable=self.sel_die_var,
                 font=("Segoe UI",10,"bold"),
                 bg=WIN_BG, fg=BLUE).pack(side="left")

        r2 = tk.Frame(sc, bg=WIN_BG); r2.pack(fill="x", padx=8, pady=(1,5))
        tk.Label(r2, text="Designs:", font=("Segoe UI",8),
                 bg=WIN_BG, fg="#666", width=13, anchor="w").pack(side="left")
        self.sel_names_var = tk.StringVar(value="—")
        tk.Label(r2, textvariable=self.sel_names_var,
                 font=("Segoe UI",8), bg=WIN_BG, fg="#333",
                 wraplength=158, justify="left").pack(side="left")

        # ── Legend ────────────────────────────────────────────────────────
        lg = ttk.LabelFrame(parent, text=" Legend ", labelanchor="nw")
        lg.pack(fill="x", padx=0, pady=(0,0))

        self._legend_null_lbl = None
        for colour, char, label in [
            ("#00aa00","1",  "Selected design → Bin 1"),
            ("#6e6e6e","X",  "Edge / drop-in die"),
            ("#f0f0f0",".",  "Outer null area"),
            ("#d2d2d2","x",  "Opted-out design"),
        ]:
            row = tk.Frame(lg, bg=WIN_BG); row.pack(fill="x", padx=8, pady=2)
            sw = tk.Canvas(row, width=22, height=16, bg=colour,
                           highlightthickness=1, highlightbackground=SEP_COL)
            sw.pack(side="left", padx=(0,7))
            sw.create_text(11,8, text=char, font=("Segoe UI",7,"bold"),
                           fill="white" if colour not in ("#f0f0f0","#d2d2d2") else "#333")
            lbl = tk.Label(row, text=label, font=("Segoe UI",8),
                           bg=WIN_BG, fg="#555")
            lbl.pack(side="left")
            if char == '.':
                self._legend_null_row = (sw, lbl)
        tk.Frame(lg, bg=WIN_BG, height=4).pack()

    # ── bottom bar ────────────────────────────────────────────────────────
    def _build_bottom_bar(self):
        bar = tk.Frame(self, bg=HDR_BG, relief="sunken", bd=1)
        bar.pack(fill="x", side="bottom")

        self.status_var = tk.StringVar(value="Ready  —  Open a wafer map file to begin")
        tk.Label(bar, textvariable=self.status_var,
                 font=("Segoe UI",8), bg=HDR_BG, fg="#444",
                 anchor="w").pack(side="left", padx=10, pady=4)

        btn_area = tk.Frame(bar, bg=HDR_BG)
        btn_area.pack(side="right", padx=8, pady=4)

        reset_btn = tk.Button(btn_area, text="Reset", command=self._reset,
                  font=("Segoe UI",9), relief="raised", bd=2,
                  bg=HDR_BG, fg="#333", padx=10, pady=3, cursor="hand2")
        reset_btn.pack(side="left", padx=(0,6))
        reset_btn.bind("<Enter>", lambda e: reset_btn.config(bg="#ffdddd", fg="#cc0000"))
        reset_btn.bind("<Leave>", lambda e: reset_btn.config(bg=HDR_BG,   fg="#333"))

        self.export_btn = tk.Button(btn_area, text="  Export...  ",
                                     command=self._export,
                                     font=("Segoe UI",9,"bold"),
                                     relief="raised", bd=2,
                                     bg=BLUE, fg="white",
                                     activebackground=BLUE_HOV,
                                     activeforeground="white",
                                     padx=12, pady=3,
                                     cursor="hand2",
                                     state="disabled",
                                     disabledforeground="#aaaaaa")
        self.export_btn.pack(side="left")
        self.export_btn.bind("<Enter>", lambda e: self.export_btn.config(bg=BLUE_HOV)
                              if str(self.export_btn.cget("state")) == "normal" else None)
        self.export_btn.bind("<Leave>", lambda e: self.export_btn.config(bg=BLUE)
                              if str(self.export_btn.cget("state")) == "normal" else None)

        self.progress = ttk.Progressbar(bar, mode="indeterminate", length=120)

    def _set_status(self, msg): self.status_var.set(msg)

    # ── open / load ───────────────────────────────────────────────────────
    def _open_file(self):
        path = filedialog.askopenfilename(
            title="Open Wafer Map",
            filetypes=[("Excel files","*.xlsx *.xls"),("All files","*.*")])
        if not path: return
        self.filepath = path
        self.breadcrumb.config(text=os.path.basename(path))
        self._set_status(f"Loading  {os.path.basename(path)} …")
        self.progress.pack(side="right", padx=6, pady=4)
        self.progress.start(10)
        threading.Thread(target=self._load_thread, daemon=True).start()

    def _load_thread(self):
        t0 = time.time()
        try:
            grid, designs, counts, sheet, fmt, null_char = \
                detect_and_parse(self.filepath)

            rows = len(grid)
            cols = max(len(r) for r in grid)

            didx = {d:i+1 for i,d in enumerate(designs)}
            flat = []
            for row in grid:
                for val in row:
                    if   val == '.':  flat.append(0)
                    elif val == 'X':  flat.append(-1)
                    elif val == 'x':  flat.append(-2)
                    else:             flat.append(didx.get(val, 0))
                flat.extend([0] * (cols - len(row)))
            enc = np.array(flat, dtype=np.int16).reshape(rows, cols)

            cm_rgb  = {d: PAL_RGB[i%len(PAL_RGB)]  for i,d in enumerate(designs)}
            cm_xlsx = {d: PAL_XLSX[i%len(PAL_XLSX)] for i,d in enumerate(designs)}

            elapsed = time.time() - t0
            fsize   = os.path.getsize(self.filepath)
            fsize_s = f"{fsize/1024:.0f} KB" if fsize<1048576 else f"{fsize/1048576:.1f} MB"

            fmt_labels = {
                'A': "Format A  (null='.', edge='X')",
                'B': "Format B  (null='-', edge='X')",
                'C': "Format C  (RowData / null='___')",
            }

            self.after(0, lambda: self._on_load_done(
                grid, enc, sheet, fmt, null_char, designs, counts,
                cm_rgb, cm_xlsx, rows, cols, fsize_s, elapsed,
                fmt_labels[fmt]))
        except Exception as e:
            import traceback
            self.after(0, lambda: self._on_load_error(str(e)+"\n"+traceback.format_exc()))

    def _on_load_done(self, grid, enc, sheet, fmt, null_char, designs, counts,
                      cm_rgb, cm_xlsx, rows, cols, fsize_s, elapsed, fmt_label):
        self.progress.stop(); self.progress.pack_forget()
        self.grid_data=grid; self.enc=enc; self.source_sheet=sheet
        self.file_fmt=fmt; self.null_char=null_char
        self.designs=designs; self.design_counts=counts
        self.cm_rgb=cm_rgb; self.cm_xlsx=cm_xlsx
        self.selected_designs.clear(); self.zoom=1.0
        self._lut = None; self._prev_lut = None; self._lut_dirty = True

        fname = os.path.basename(self.filepath)
        self.stat_vars["File:"].set(fname)
        self.stat_vars["Format:"].set(fmt_label)
        self.stat_vars["Dims:"].set(f"{rows} × {cols}")
        self.stat_vars["Total Die:"].set(f"{sum(counts.values()):,}")
        self.stat_vars["Load Time:"].set(f"{elapsed:.2f}s")

        self.breadcrumb.config(text=f"{fname}  [{sheet}]  ·  {fmt_label}")
        self._set_status(
            f"✔  {fname}  ·  {len(designs)} designs detected  ·  {rows}×{cols}  ·  {fmt_label}")

        self._populate_designs()
        self._render_map()
        self.after(100, self._render_preview)

    def _on_load_error(self, msg):
        self.progress.stop(); self.progress.pack_forget()
        self._set_status(f"Error loading file")
        messagebox.showerror("Load Error", msg)

    # ── design list ───────────────────────────────────────────────────────
    def _populate_designs(self):
        for w in self.design_inner.winfo_children(): w.destroy()
        self.design_vars   = {}
        self._hover_jobs   = {}   # row_id → after-job for animation

        # Colours used by the hover animation
        HOV_BG   = (220, 234, 255)   # target hover bg  RGB
        NORM_BG  = tuple(int(WIN_BG.lstrip('#')[i:i+2],16) for i in (0,2,4))
        HOV_FG   = "#003070"
        NORM_FG  = "#111111"
        HOV_CNT  = "#0055cc"
        NORM_CNT = "#555555"
        STEPS    = 6          # animation frames
        DELAY    = 12         # ms between frames

        def _lerp_col(a, b, t):
            """Linearly interpolate between two RGB tuples; return hex string."""
            r = int(a[0] + (b[0]-a[0])*t)
            g = int(a[1] + (b[1]-a[1])*t)
            bl= int(a[2] + (b[2]-a[2])*t)
            return f"#{r:02x}{g:02x}{bl:02x}"

        def _animate(row_id, row_frame, cb_widget, cnt_lbl, step, going_in):
            """Animate one frame of the hover fade."""
            t = step / STEPS
            if not going_in:
                t = 1.0 - t
            bg_hex = _lerp_col(NORM_BG, HOV_BG, t)
            # left accent strip colour
            acc_hex = _lerp_col(NORM_BG, (0, 103, 192), t)

            try:
                row_frame.config(bg=bg_hex)
                cb_widget.config(bg=bg_hex, activebackground=bg_hex,
                                  selectcolor=bg_hex,
                                  fg=_lerp_col(
                                      tuple(int(NORM_FG.lstrip('#')[i:i+2],16) for i in (0,2,4)),
                                      tuple(int(HOV_FG.lstrip('#')[i:i+2],16) for i in (0,2,4)),
                                      t))
                cnt_lbl.config(bg=bg_hex,
                                fg=_lerp_col(
                                    tuple(int(NORM_CNT.lstrip('#')[i:i+2],16) for i in (0,2,4)),
                                    tuple(int(HOV_CNT.lstrip('#')[i:i+2],16) for i in (0,2,4)),
                                    t))
                # accent bar
                acc_bar = row_frame._acc_bar
                acc_bar.config(bg=acc_hex)
            except tk.TclError:
                return   # widget was destroyed

            if step < STEPS:
                job = row_frame.after(DELAY,
                    lambda: _animate(row_id, row_frame, cb_widget,
                                     cnt_lbl, step+1, going_in))
                self._hover_jobs[row_id] = job

        for i, design in enumerate(self.designs):
            hex_col = "#%02x%02x%02x" % PAL_RGB[i%len(PAL_RGB)]
            var = tk.BooleanVar(value=False)
            self.design_vars[design] = var
            count = self.design_counts.get(design, 0)

            # Outer row with a thin left-accent bar
            row = tk.Frame(self.design_inner, bg=WIN_BG, cursor="hand2")
            row.pack(fill="x", padx=0, pady=0)

            acc = tk.Frame(row, width=3, bg=WIN_BG, cursor="hand2")
            acc.pack(side="left", fill="y")
            row._acc_bar = acc   # stash reference

            sw = tk.Canvas(row, width=14, height=13, bg=hex_col,
                           highlightthickness=1, highlightbackground=SEP_COL,
                           cursor="hand2")
            sw.pack(side="left", padx=(5,0), pady=5)

            cb = tk.Checkbutton(row,
                                 text=f"  {str(design):<8}",
                                 font=("Segoe UI",9), variable=var,
                                 bg=WIN_BG, activebackground=WIN_BG,
                                 selectcolor=WIN_BG, fg="#111",
                                 command=self._on_sel_change,
                                 relief="flat", bd=0, cursor="hand2")
            cb.pack(side="left")

            cnt = tk.Label(row, text=f"{count:>9,}",
                           font=("Courier New",8), bg=WIN_BG, fg="#555")
            cnt.pack(side="right", padx=8)

            row_id = id(row)

            def _enter(e, rid=row_id, r=row, c=cb, lbl=cnt):
                if rid in self._hover_jobs:
                    r.after_cancel(self._hover_jobs.pop(rid))
                _animate(rid, r, c, lbl, 1, True)

            def _leave(e, rid=row_id, r=row, c=cb, lbl=cnt):
                if rid in self._hover_jobs:
                    r.after_cancel(self._hover_jobs.pop(rid))
                _animate(rid, r, c, lbl, 1, False)

            for widget in (row, acc, sw, cb, cnt):
                widget.bind("<Enter>", _enter)
                widget.bind("<Leave>", _leave)

            # Only bind click on non-checkbox widgets to avoid double-toggle
            # (the Checkbutton handles its own click via command=)
            for widget in (row, acc, sw, cnt):
                widget.bind("<Button-1>",
                    lambda e, v=var: [v.set(not v.get()), self._on_sel_change()])

        self.design_inner.update_idletasks()
        self._design_canvas.configure(scrollregion=self._design_canvas.bbox("all"))

    def _on_sel_change(self):
        self.selected_designs = {d for d,v in self.design_vars.items() if v.get()}
        n   = len(self.selected_designs)
        die = sum(self.design_counts.get(d,0) for d in self.selected_designs)

        self.sel_count_lbl.config(text=f"{n} selected")
        self.sel_die_var.set(f"{die:,}" if n else "—")
        self.sel_names_var.set(", ".join(sorted(self.selected_designs)) if n else "—")

        en = "normal" if n else "disabled"
        # Bottom Export button
        self.export_btn.config(state=en,
                                bg=BLUE if n else WIN_BG,
                                fg="white" if n else "#aaaaaa")
        # Toolbar Export button — restore hover bindings when enabled
        self.tb_export.config(state=en,
                               fg="#111" if n else "#aaaaaa",
                               bg=WIN_BG, relief="flat")
        if n:
            self.tb_export.bind("<Enter>",
                lambda e: self.tb_export.config(bg="#d9e8f5", fg="#0055aa", relief="groove"))
            self.tb_export.bind("<Leave>",
                lambda e: self.tb_export.config(bg=WIN_BG, fg="#111", relief="flat"))
        else:
            self.tb_export.unbind("<Enter>")
            self.tb_export.unbind("<Leave>")

        self._lut_dirty = True
        self._schedule_redraw()
        if self.grid_data:
            self._render_preview()

    def _select_all(self):
        for v in self.design_vars.values(): v.set(True)
        self._on_sel_change()

    def _select_none(self):
        for v in self.design_vars.values(): v.set(False)
        self._on_sel_change()

    # ── rendering ─────────────────────────────────────────────────────────
    def _rebuild_luts(self):
        """Build colour lookup tables for map and preview."""
        n   = len(self.designs)
        bg  = tuple(int(WIN_BG.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
        GREY = (190, 190, 190)

        lut = np.zeros((n + 3, 3), dtype=np.uint8)
        lut[0] = x_RGB
        lut[1] = X_RGB
        lut[2] = bg
        for i, d in enumerate(self.designs):
            lut[3 + i] = SEL_RGB if d in self.selected_designs \
                         else self.cm_rgb.get(d, (160, 160, 160))
        self._lut = lut

        plut = np.zeros((n + 3, 3), dtype=np.uint8)
        plut[0] = x_RGB
        plut[1] = X_RGB
        plut[2] = bg
        for i, d in enumerate(self.designs):
            plut[3 + i] = SEL_RGB if d in self.selected_designs else GREY
        self._prev_lut = plut

        self._lut_dirty = False

    def _build_pil(self):
        if self._lut_dirty or self._lut is None:
            self._rebuild_luts()
        idx = (self.enc + 2).clip(0, self._lut.shape[0] - 1)
        return Image.fromarray(self._lut[idx], 'RGB')

    def _render_map(self):
        if self.enc is None: return
        cw = max(self.canvas.winfo_width(),  300)
        ch = max(self.canvas.winfo_height(), 300)
        h, w = self.enc.shape

        base_scale = min(cw / w, ch / h)
        dw = max(int(w * base_scale * self.zoom), 1)
        dh = max(int(h * base_scale * self.zoom), 1)
        dw = min(dw, w * 12)
        dh = min(dh, h * 12)

        pil = self._build_pil().resize((dw, dh), Image.NEAREST)

        if self.show_grid.get() and (dw / w) >= 5:
            from PIL import ImageDraw
            draw = ImageDraw.Draw(pil)
            sx = int(dw / w); sy = int(dh / h)
            for x in range(0, dw, sx):
                draw.line([(x, 0), (x, dh)], fill=(200, 200, 200), width=1)
            for y in range(0, dh, sy):
                draw.line([(0, y), (dw, y)], fill=(200, 200, 200), width=1)

        self._photo = ImageTk.PhotoImage(pil)
        self.canvas.delete("all")
        self.canvas.configure(scrollregion=(0, 0, dw, dh), bg="#e8e8e8")
        self.canvas.create_image(0, 0, anchor="nw", image=self._photo)
        self.zoom_lbl.config(text=f"Zoom: {int(self.zoom * 100)}%")

    def _schedule_redraw(self):
        if self._pending: return
        self._pending=True; self.after(40, self._flush_redraw)
    def _flush_redraw(self):
        self._pending=False
        if self.enc is not None:
            self._render_map()
            self._render_preview()

    def _do_convert(self):
        if not self.grid_data:
            messagebox.showinfo("No File","Open a file first."); return
        self._render_map()
        self._render_preview()
        self._set_status("Converted  —  select design(s) then click Export")

    # ── preview canvas ────────────────────────────────────────────────────
    def _build_prev_pil(self):
        if self._lut_dirty or self._prev_lut is None:
            self._rebuild_luts()
        idx = (self.enc + 2).clip(0, self._prev_lut.shape[0] - 1)
        return Image.fromarray(self._prev_lut[idx], 'RGB')

    def _render_preview(self):
        if self.enc is None: return
        cw = max(self.prev_canvas.winfo_width(),  300)
        ch = max(self.prev_canvas.winfo_height(), 300)
        h, w = self.enc.shape

        base_scale = min(cw / w, ch / h)
        dw = max(int(w * base_scale * self.prev_zoom), 1)
        dh = max(int(h * base_scale * self.prev_zoom), 1)
        dw = min(dw, w * 12); dh = min(dh, h * 12)

        pil = self._build_prev_pil().resize((dw, dh), Image.NEAREST)

        if self.show_grid.get() and (dw / w) >= 5:
            from PIL import ImageDraw
            draw = ImageDraw.Draw(pil)
            sx = int(dw / w); sy = int(dh / h)
            for x in range(0, dw, sx):
                draw.line([(x, 0), (x, dh)], fill=(200, 200, 200), width=1)
            for y in range(0, dh, sy):
                draw.line([(0, y), (dw, y)], fill=(200, 200, 200), width=1)

        self._prev_photo = ImageTk.PhotoImage(pil)
        self.prev_canvas.delete("all")
        self.prev_canvas.configure(scrollregion=(0, 0, dw, dh), bg="#e8e8e8")
        self.prev_canvas.create_image(0, 0, anchor="nw", image=self._prev_photo)

    def _draw_prev_placeholder(self):
        self.prev_canvas.delete("all")
        self.prev_canvas.configure(bg="#e8e8e8")
        self.prev_canvas.create_text(200, 150,
            text="Select a design to preview",
            fill="#aaaaaa", font=("Segoe UI", 11))

    def _on_prev_scroll(self, event):
        if event.state & 0x4:
            self._on_prev_zoom(event); return
        if event.num == 4:
            self.prev_canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.prev_canvas.yview_scroll(1, "units")
        else:
            self.prev_canvas.yview_scroll(-1*(event.delta//120), "units")

    def _on_prev_zoom(self, event):
        if event.num == 4 or (hasattr(event,'delta') and event.delta > 0):
            factor = 1.15
        else:
            factor = 1/1.15
        old = self.prev_zoom
        self.prev_zoom = max(0.05, min(16.0, self.prev_zoom * factor))
        if abs(self.prev_zoom - old) < 0.001: return
        ratio = self.prev_zoom / old
        if self.enc is not None:
            cx = self.prev_canvas.canvasx(event.x)
            cy = self.prev_canvas.canvasy(event.y)
            self._render_preview()
            def _adj():
                sr = self.prev_canvas.cget("scrollregion")
                if not sr: return
                _, _, sw, sh = map(float, sr.split())
                fx = max(0.0, min(1.0, (cx*ratio - event.x) / sw))
                fy = max(0.0, min(1.0, (cy*ratio - event.y) / sh))
                self.prev_canvas.xview_moveto(fx)
                self.prev_canvas.yview_moveto(fy)
            self.after(60, _adj)

    def _draw_placeholder(self):
        self.canvas.delete("all")
        self.canvas.configure(bg=WIN_BG)
        self.canvas.create_text(400,300,
            text="Open a wafer map file to begin",
            fill="#aaaaaa", font=("Segoe UI",12))

    # ── zoom ─────────────────────────────────────────────────────────────
    def _zoom_in(self):    self.zoom=min(self.zoom*1.5,16.); self._schedule_redraw()
    def _zoom_out(self):   self.zoom=max(self.zoom/1.5,.05); self._schedule_redraw()
    def _zoom_fit(self):   self.zoom=1.0; self._schedule_redraw()

    # ── export ────────────────────────────────────────────────────────────
    def _export(self):
        if not self.selected_designs:
            messagebox.showwarning("No Selection",
                "Select at least one design first."); return
        die = sum(self.design_counts.get(d,0) for d in self.selected_designs)
        dlg = ExportDialog(self, sorted(self.selected_designs), die)
        self.wait_window(dlg)
        if not dlg.result: return

        opts  = dlg.result
        label = "_".join(sorted(self.selected_designs))
        if opts["fmt"]=="xlsx":
            path = filedialog.asksaveasfilename(
                title="Save Excel Wafer Map", defaultextension=".xlsx",
                initialfile=f"Design_{label}_R1.00.xlsx",
                filetypes=[("Excel Workbook","*.xlsx"),("All files","*.*")])
        else:
            path = filedialog.asksaveasfilename(
                title="Save Text Wafer Map", defaultextension=".txt",
                initialfile=f"Design_{label}_R1.00.txt",
                filetypes=[("Text files","*.txt"),("All files","*.*")])
        if not path: return

        self._set_status("Exporting …")
        self.progress.pack(side="right", padx=6, pady=4); self.progress.start(10)
        if opts["fmt"]=="xlsx":
            threading.Thread(target=self._export_xlsx, args=(path,), daemon=True).start()
        else:
            threading.Thread(target=self._export_txt,  args=(path,opts["le"]), daemon=True).start()

    def _export_xlsx(self, path):
        try:
            sel=self.selected_designs
            FILL_SEL=PatternFill("solid",fgColor=SEL_XLSX)
            FILL_X  =PatternFill("solid",fgColor=X_XLSX)
            wb=openpyxl.Workbook(write_only=True)
            ws=wb.create_sheet(self.source_sheet or "MAP")
            die_count=0
            for row in self.grid_data:
                out=[]
                for val in row:
                    c=WriteOnlyCell(ws)
                    if val=='.':        c.value='.'
                    elif val=='X':      c.value='X'; c.fill=FILL_X
                    elif val in sel:    c.value=1; c.fill=FILL_SEL; die_count+=1
                    else:               c.value='x'
                    out.append(c)
                ws.append(out)
            ws2=wb.create_sheet("Summary")
            ws2.append(["Design","Die Count","Exported"])
            for d in self.designs:
                ws2.append([d,self.design_counts.get(d,0),
                             "✔ Bin 1" if d in sel else ""])
            ws2.append([]); ws2.append(["TOTAL EXPORTED",die_count,""])
            wb.save(path)
            self.after(0,lambda: self._on_done(path,die_count,"xlsx"))
        except Exception as e:
            self.after(0,lambda: self._on_err(str(e)))

    def _export_txt(self, path, le):
        try:
            sel=self.selected_designs; nl="\r\n" if le=="crlf" else "\n"
            lines=[]; die_count=0
            for row in self.grid_data:
                parts=[]
                for val in row:
                    if val=='.':      parts.append('.')
                    elif val=='X':    parts.append('X')
                    elif val in sel:  parts.append('1'); die_count+=1
                    else:             parts.append('x')
                lines.append(''.join(parts))
            with open(path,'w',encoding='utf-8',newline='') as f:
                f.write(nl.join(lines))
            self.after(0,lambda: self._on_done(path,die_count,"txt"))
        except Exception as e:
            self.after(0,lambda: self._on_err(str(e)))

    def _on_done(self, path, die_count, fmt):
        self.progress.stop(); self.progress.pack_forget()
        fname=os.path.basename(path)
        self._set_status(f"✔  Exported  {fname}  ·  {die_count:,} die as Bin 1")
        if fmt=="xlsx":
            messagebox.showinfo("Export Complete",
                f"Saved:  {fname}\n\nDie as Bin 1:  {die_count:,}\n\n"
                f"Sheets:\n  • {self.source_sheet or 'MAP'} — wafer grid\n"
                f"  • Summary — design counts")
        else:
            messagebox.showinfo("Export Complete",
                f"Saved:  {fname}\n\nDie as Bin 1:  {die_count:,}\n\n"
                f"Next: add your site header/footer and save as Unicode UTF-8.")

    def _on_err(self, msg):
        self.progress.stop(); self.progress.pack_forget()
        self._set_status(f"Export error: {msg}")
        messagebox.showerror("Export Error", msg)

    def _reset(self):
        if not messagebox.askyesno(
                "Reset",
                "Clear the current file and all selections?\n\nThis cannot be undone.",
                icon="warning"):
            return
        self.filepath=None; self.grid_data=[]; self.enc=None
        self.designs=[]; self.design_counts={}
        self.selected_designs.clear(); self.design_vars={}; self._hover_jobs={}; self.zoom=1.0; self.prev_zoom=1.0
        self._lut=None; self._prev_lut=None; self._lut_dirty=True
        self.breadcrumb.config(text="No file loaded")
        for w in self.design_inner.winfo_children(): w.destroy()
        tk.Label(self.design_inner, text="Open a file to detect designs",
                 font=("Segoe UI",8,"italic"),
                 bg=WIN_BG, fg="#999").pack(pady=14)
        for k in self.stat_vars: self.stat_vars[k].set("—")
        self.sel_die_var.set("—"); self.sel_names_var.set("—")
        self.sel_count_lbl.config(text="0 selected")
        self.export_btn.config(state="disabled", bg=WIN_BG, fg="#aaaaaa")
        self.tb_export.config(state="disabled", fg="#aaaaaa")
        self._draw_prev_placeholder()
        self._draw_placeholder()
        self._set_status("Ready  —  Open a wafer map file to begin")

    def _on_resize(self, e=None):
        if self.enc is not None:
            self.after(120, self._flush_redraw)
            self.after(150, self._render_preview)


# ═══════════════════════════════════════════════════════════════════════
#  EXPORT DIALOG
# ═══════════════════════════════════════════════════════════════════════
class ExportDialog(tk.Toplevel):
    def __init__(self, parent, designs, die_count):
        super().__init__(parent)
        self.title("Export")
        self.resizable(False, False)
        self.grab_set()
        self.result = None
        self.configure(bg=WIN_BG)
        x = parent.winfo_x() + parent.winfo_width()//2 - 210
        y = parent.winfo_y() + parent.winfo_height()//2 - 175
        self.geometry(f"420x370+{x}+{y}")

        # blue header strip
        hdr = tk.Frame(self, bg=BLUE)
        hdr.pack(fill="x")
        tk.Label(hdr, text="  Export Wafer Map",
                 font=("Segoe UI",11,"bold"),
                 bg=BLUE, fg="white", pady=9).pack(side="left")

        body = tk.Frame(self, bg=WIN_BG)
        body.pack(fill="both", padx=18, pady=10)

        # info box
        ib = tk.Frame(body, bg="#deeaf7", relief="solid", bd=1)
        ib.pack(fill="x", pady=(0,12))
        tk.Label(ib, text=f"  Exporting:  {', '.join(designs)}",
                 font=("Segoe UI",8,"bold"), bg="#deeaf7",
                 anchor="w").pack(fill="x", padx=6, pady=(5,1))
        tk.Label(ib, text=f"  Die → Bin 1:  {die_count:,}",
                 font=("Segoe UI",8), bg="#deeaf7", fg="#555",
                 anchor="w").pack(fill="x", padx=6, pady=(1,5))

        tk.Label(body, text="Format", font=("Segoe UI",9,"bold"),
                 bg=WIN_BG, fg="#111").pack(anchor="w")
        tk.Frame(body, bg=SEP_COL, height=1).pack(fill="x", pady=(2,8))

        self.fmt = tk.StringVar(value="xlsx")
        self.fmt.trace_add("write", self._toggle)

        ef = tk.Frame(body, bg=WIN_BG); ef.pack(fill="x", pady=3)
        tk.Radiobutton(ef, text="Excel Workbook  (.xlsx)",
                       variable=self.fmt, value="xlsx",
                       font=("Segoe UI",9,"bold"),
                       bg=WIN_BG, activebackground=WIN_BG,
                       fg="#111", selectcolor=WIN_BG).pack(anchor="w")
        tk.Label(ef, text="   Colour-coded map sheet + Summary sheet with die counts",
                 font=("Segoe UI",8), bg=WIN_BG, fg="#666").pack(anchor="w")

        tf = tk.Frame(body, bg=WIN_BG); tf.pack(fill="x", pady=(8,2))
        tk.Radiobutton(tf, text="Unicode Text  (.txt)",
                       variable=self.fmt, value="txt",
                       font=("Segoe UI",9,"bold"),
                       bg=WIN_BG, activebackground=WIN_BG,
                       fg="#111", selectcolor=WIN_BG).pack(anchor="w")
        tk.Label(tf, text="   Plain character map for assembly site upload",
                 font=("Segoe UI",8), bg=WIN_BG, fg="#666").pack(anchor="w")

        self.le_frame = tk.Frame(body, bg=WIN_BG)
        self.le_frame.pack(anchor="w", padx=18, pady=(2,0))
        tk.Label(self.le_frame, text="Line ending:",
                 font=("Segoe UI",8), bg=WIN_BG, fg="#555").pack(side="left")
        self.le = tk.StringVar(value="crlf")
        for txt, val in [("CRLF (Windows)","crlf"),("LF (Unix)","lf")]:
            tk.Radiobutton(self.le_frame, text=txt, variable=self.le, value=val,
                           font=("Segoe UI",8), bg=WIN_BG,
                           activebackground=WIN_BG,
                           fg="#111", selectcolor=WIN_BG).pack(side="left", padx=(6,0))

        tk.Frame(body, bg=SEP_COL, height=1).pack(fill="x", pady=12)

        br = tk.Frame(body, bg=WIN_BG); br.pack(fill="x")
        tk.Button(br, text="Cancel", command=self.destroy,
                  font=("Segoe UI",9), relief="raised", bd=2,
                  bg=WIN_BG, padx=10, pady=4,
                  cursor="hand2").pack(side="right", padx=(6,0))
        tk.Button(br, text="  Export...  ", command=self._ok,
                  font=("Segoe UI",9,"bold"), relief="raised", bd=2,
                  bg=BLUE, fg="white",
                  activebackground=BLUE_HOV, activeforeground="white",
                  padx=10, pady=4,
                  cursor="hand2").pack(side="right")

        self._toggle()
        self.bind("<Return>", lambda e: self._ok())
        self.bind("<Escape>", lambda e: self.destroy())

    def _toggle(self, *_):
        is_txt = self.fmt.get()=="txt"
        for w in self.le_frame.winfo_children():
            try: w.config(state="normal" if is_txt else "disabled")
            except: pass

    def _ok(self):
        self.result = {"fmt":self.fmt.get(), "le":self.le.get()}
        self.destroy()


# ── run ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = WaferMapTool()
    app.bind("<Configure>", app._on_resize)
    app.mainloop()